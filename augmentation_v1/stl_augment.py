"""
STL Augmentation Script
-----------------------
Scales all .stl files from ./models/ folder by each factor in SCALE_FACTORS.
Saves results to ./models_augmented/ and generates augmentation_report.xlsx.
"""

import os
import struct
import shutil
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── KONFIGURACE ────────────────────────────────────────────────────────────
SCALE_FACTORS = [0.8, 0.9, 1.1, 1.2]          # <- sem přidej / odeber faktory dle potřeby

# Cesty jsou vždy relativní ke složce kde leží tento skript
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR    = os.path.join(_SCRIPT_DIR, "models")
OUTPUT_DIR   = os.path.join(_SCRIPT_DIR, "models_augmented")
REPORT_FILE  = "augmentation_report.xlsx"
# ─────────────────────────────────────────────────────────────────────────────


def read_stl_binary(path):
    """Načte binární STL → (header bytes, list of trojúhelníků)."""
    with open(path, "rb") as f:
        header    = f.read(80)
        num_tris  = struct.unpack("<I", f.read(4))[0]
        triangles = []
        for _ in range(num_tris):
            data = f.read(50)   # 12B normal + 3×12B vertex + 2B attr
            triangles.append(data)
    return header, triangles


def write_stl_binary(path, header, triangles):
    """Zapíše binární STL ze seznamu trojúhelníků."""
    with open(path, "wb") as f:
        f.write(header)
        f.write(struct.pack("<I", len(triangles)))
        for tri in triangles:
            f.write(tri)


def scale_triangle(tri_bytes, factor):
    """Přeškáluje jeden trojúhelník (normal zůstane beze změny, vertexy se násobí)."""
    f = float(factor)
    # normal (3 floats) + 3 vertices (každý 3 floaty) + attr (2 bytes)
    floats = struct.unpack("<12f", tri_bytes[:48])
    attr   = tri_bytes[48:50]

    normal   = floats[0:3]          # normála se nemění
    v1       = tuple(x * f for x in floats[3:6])
    v2       = tuple(x * f for x in floats[6:9])
    v3       = tuple(x * f for x in floats[9:12])

    return struct.pack("<12f", *normal, *v1, *v2, *v3) + attr


def is_binary_stl(path):
    """Heuristika: zkusíme přečíst jako binární a ověříme konzistenci velikosti."""
    try:
        size = os.path.getsize(path)
        with open(path, "rb") as f:
            f.read(80)
            num_tris = struct.unpack("<I", f.read(4))[0]
        expected = 80 + 4 + num_tris * 50
        return size == expected and num_tris > 0
    except Exception:
        return False


def scale_stl_ascii(src_path, dst_path, factor):
    """Přeškáluje ASCII STL (pomalejší, ale univerzální záloha)."""
    f = float(factor)
    with open(src_path, "r") as fin, open(dst_path, "w") as fout:
        for line in fin:
            stripped = line.strip()
            if stripped.startswith("vertex "):
                parts = stripped.split()
                coords = [str(float(parts[i]) * f) for i in range(1, 4)]
                indent = line[: len(line) - len(line.lstrip())]
                fout.write(f"{indent}vertex {' '.join(coords)}\n")
            else:
                fout.write(line)


def scale_stl(src_path, dst_path, factor):
    if is_binary_stl(src_path):
        header, triangles = read_stl_binary(src_path)
        scaled = [scale_triangle(t, factor) for t in triangles]
        write_stl_binary(dst_path, header, scaled)
    else:
        scale_stl_ascii(src_path, dst_path, factor)


def create_report(records, output_path):
    """Vytvoří Excel report: sloupec A = nový soubor, sloupec B = původní soubor."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Augmentace"

    header_fill   = PatternFill("solid", start_color="2E4057", end_color="2E4057")
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font     = Font(name="Arial", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill      = PatternFill("solid", start_color="F2F4F8", end_color="F2F4F8")

    headers = ["Augmentovaný soubor", "Původní soubor"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    for row_idx, (aug_name, orig_name) in enumerate(records, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, val in enumerate([aug_name, orig_name], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = cell_font
            cell.alignment = left
            cell.border    = border
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 35
    ws.row_dimensions[1].height     = 22

    # Celkový počet – informativní řádek pod daty
    summary_row = len(records) + 2
    ws.cell(row=summary_row, column=1,
            value=f"Celkem augmentovaných souborů: {len(records)}"
            ).font = Font(name="Arial", italic=True, size=10, color="666666")

    wb.save(output_path)


def main():
    if not os.path.isdir(INPUT_DIR):
        raise FileNotFoundError(f"Složka '{INPUT_DIR}' nebyla nalezena vedle skriptu.")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    stl_files = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith(".stl")]
    if not stl_files:
        print(f"[!] Žádné .stl soubory ve složce '{INPUT_DIR}'.")
        return

    records = []
    total   = len(stl_files) * len(SCALE_FACTORS)
    done    = 0

    for filename in sorted(stl_files):
        base, ext = os.path.splitext(filename)
        src_path  = os.path.join(INPUT_DIR, filename)

        for factor in SCALE_FACTORS:
            # Formátování faktoru: 2 → "2", 0.5 → "0.5", 1.25 → "1.25"
            factor_str = str(factor).rstrip("0").rstrip(".") if "." in str(factor) else str(factor)
            new_name   = f"{base}_{factor_str}x{ext}"
            dst_path   = os.path.join(OUTPUT_DIR, new_name)

            scale_stl(src_path, dst_path, factor)
            records.append((new_name, filename))
            done += 1
            print(f"[{done}/{total}] {filename}  →  {new_name}  (×{factor})")

    report_path = os.path.join(OUTPUT_DIR, REPORT_FILE)
    create_report(records, report_path)
    print(f"\n✓ Hotovo! {len(records)} souborů uloženo do '{OUTPUT_DIR}/'")
    print(f"✓ Report: {report_path}")


if __name__ == "__main__":
    main()
